#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/7/5 15:40
# @Author  :  Leanper

# @File    : setStyle.py
# @Software: PyCharm
# @Desc:

import openpyxl
from openpyxl.styles import Font, Alignment, colors, PatternFill, Side, Border, GradientFill

wb = openpyxl.load_workbook('studyexcel.xlsx')
sheet = wb["Sheet"]

print(sheet.max_row)
print(sheet.max_column)

# textFont 设置
text_font = Font(name="楷体", color="FF0000", italic=True)
sheet["B4"].font = text_font


# alignment 参数 水平垂直居中 文本旋转角度  wrap_text=是否自动换行
# 水平对齐：
# horizontal代表水平方向，可以左对齐left，还有居中center和右对齐right，分散对齐distributed，跨列居中centerContinuous，两端对齐justify，填充fill，常规general
# 垂直对齐：
# vertical代表垂直方向，可以居中center，还可以靠上top，靠下bottom，两端对齐justify，分散对齐distributed
alignment = Alignment(horizontal="center", vertical="center", text_rotation=45, wrap_text=True)
sheet["C2"].alignment = alignment


# 设置边框样式 Side(style，color)
#  style 参数的种类： 'double, 'mediumDashDotDot', 'slantDashDot',
# 'dashDotDot','dotted','hair', 'mediumDashed, 'dashed', 'dashDot', 'thin',
# 'mediumDashDot','medium', 'thick'

side1 = Side(style="thin", color="FF0000")
side2 = Side(style="thick", color="FFFF0000")
border = Border(left=side1, right=side1, top=side2, bottom=side2)
sheet["E1"].border = border

# 背景设置
# PatternFill(fill_type=填充样式，fgColor=填充颜色）
# GradientFill(stop=(渐变颜色 1，渐变颜色 2……))
pattern_fill = PatternFill(fill_type="solid", fgColor="99ccff")
sheet["E2"].fill = pattern_fill
gradient_fill = GradientFill(stop=("FFFFFF", "99ccff", "000000"))
sheet["E3"].fill = gradient_fill

# 合并单元格
sheet = wb.active
sheet.merged_cells("E3:E4")
sheet.merge_cells(start_row=2, start_column=2, end_column=3, end_row=3)

wb.save("studyexcel.xlsx")

for cell in sheet['D2:D5']:
    fill = PatternFill('darkGray', bgColor=colors.BLUE)
    if int(cell[0].value) > 60:
        # sheet.cell(cell, "B3").fill = fill
        wb.save("studyexcel.xlsx")

# for cell in range(1, sheet.max_column):
#     if int(cell[0].value) > 60:
#         fill = PatternFill('darkGray', bgColor=colors.BLUE)
#     wb.save("studyexcel.xlsx")

