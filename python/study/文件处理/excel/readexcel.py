#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/7/1 15:57
# @Author  :  Leanper

# @File    : readexcel.py
# @Software: PyCharm
# @Desc:

import openpyxl

wb = openpyxl.load_workbook('studyexcel.xlsx')
sheet = wb['Sheet']
print(sheet["B2:B5"])
print(sheet["B2"].value)

# 获取制定数据
for cell in sheet["B2:B5"]:
    print(cell[0].value)

for cell in sheet:
    for row in cell:
        print(row.value, end=',')
    print()

# 制定几行
print("按列循环--制定几行")
for cell in sheet.iter_rows(min_col=2, max_col=4, max_row=3):
    for c in cell:
        print(c.value)



print("按列循环")
# 按列循环
for cell in sheet.columns:
    for c in cell:
        print(c.value)
    print()


# 制定几行
print("按列循环--制定几行")
for cell in sheet.iter_cols(min_col=2, max_col=4, max_row=3):
    for c in cell:
        print(c.value)


# 修改内容
sheet["B3"] = "修改了"
sheet['B8'] = "李四"
sheet.append([5, '张三', '0', 90, '二班'])
wb.save("studyexcel.xlsx")

# 删除