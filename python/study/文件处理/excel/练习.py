#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/7/5 17:11
# @Author  :  Leanper

# @File    : 练习.py
# @Software: PyCharm
# @Desc:

import openpyxl
wb = openpyxl.load_workbook('studyexcel.xlsx')
sheet = wb.create_sheet("data3")
for x in range(1, 26):
    for y in range(1, 26):
        if x == y:
            sheet.cell(x, y).value = str(x)+","+str(y)
        else:
            sheet.cell(x, y).value = str(x)+","+str(y)
wb.save("studyexcel.xlsx")

print(sheet["C3:F5"][0])


wb.close()
